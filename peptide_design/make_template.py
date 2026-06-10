#!/usr/bin/env python3
# peptide_design/make_template.py
"""
Generate the Excel design template as a .xlsx file.

Usage:
    python -m peptide_design.make_template
    python -m peptide_design.make_template --out /path/to/file.xlsx
"""
from __future__ import annotations
import argparse
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

MAIN_CHAIN_POSITIONS = 34
MAX_SC_MONOMERS = 8
MAX_SIDECHAINS = 2

BLUE_FILL   = PatternFill("solid", fgColor="D9E1F2")
GREEN_FILL  = PatternFill("solid", fgColor="E2EFDA")
GREY_FILL   = PatternFill("solid", fgColor="F2F2F2")
ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")


def _header(ws, col, row, value, fill=None, bold=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold)
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(horizontal="center")
    return cell


def _build_design_headers(ws) -> list[str]:
    col = 1
    headers = []

    _header(ws, col, 1, "Name", GREY_FILL)
    ws.column_dimensions[get_column_letter(col)].width = 22
    headers.append("Name")
    col += 1

    for pos in range(1, MAIN_CHAIN_POSITIONS + 1):
        _header(ws, col, 1, str(pos), BLUE_FILL)
        ws.column_dimensions[get_column_letter(col)].width = 6
        headers.append(str(pos))
        col += 1

    for sc_idx in range(1, MAX_SIDECHAINS + 1):
        suffix = "" if sc_idx == 1 else f"_{sc_idx}"
        site_key = f"Site{suffix}"
        _header(ws, col, 1, site_key, ORANGE_FILL)
        ws.column_dimensions[get_column_letter(col)].width = 8
        headers.append(site_key)
        col += 1

        for k in range(1, MAX_SC_MONOMERS + 1):
            bk = f"b{k}{suffix}"
            sk = f"SC{k}{suffix}"
            _header(ws, col, 1, bk, GREEN_FILL)
            ws.column_dimensions[get_column_letter(col)].width = 8
            headers.append(bk)
            col += 1
            _header(ws, col, 1, sk, GREEN_FILL)
            ws.column_dimensions[get_column_letter(col)].width = 8
            headers.append(sk)
            col += 1

    ws.freeze_panes = "B2"
    return headers


def _write_example_rows(ws, headers):
    PYY_SEQ = list("IKPEAPGEDASPEE") + list("LNRYYASLRHYLNL") + list("VTRQRY")

    def row_dict(name, backbone=None, sc1=None):
        d = {"Name": name}
        seq = backbone or PYY_SEQ
        for i, aa in enumerate(seq, 1):
            d[str(i)] = aa
        if sc1:
            site, bonds, monomers = sc1
            d["Site"] = site
            for k, (b, m) in enumerate(zip(bonds, monomers), 1):
                d[f"b{k}"] = b
                d[f"SC{k}"] = m
        return d

    backbone_k30 = list(PYY_SEQ); backbone_k30[29] = "K"

    examples = [
        row_dict("PYY3-36-ref"),
        row_dict("Analog_28-pos30-C18", backbone_k30,
                 sc1=(30, ["R3-R1","R2-R1","R2-R1","R2-R1"], ["gGlu","Ado","Ado","C18d"])),
        row_dict("Analog_37-pos30-C14d-2xAdo", backbone_k30,
                 sc1=(30, ["R3-R1","R2-R1","R2-R1"], ["Ado","Ado","C14d"])),
        row_dict("Analog_01-posNa-N-term",
                 sc1=(1, ["R1-R1","R2-R1","R2-R1","R2-R1"], ["gGlu","Ado","Ado","C18d"])),
    ]

    for row_num, ex in enumerate(examples, start=2):
        for col_num, h in enumerate(headers, start=1):
            val = ex.get(h, "")
            if val:
                ws.cell(row=row_num, column=col_num, value=str(val))


def _write_monomer_db_sheet(wb):
    ws = wb.create_sheet("MonomerDB")
    _header(ws, 1, 1, "Symbol", GREY_FILL)
    _header(ws, 2, 1, "PolymerType", GREY_FILL)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14

    root = Path(__file__).parent.parent / "monomer_db"
    entries = []
    for fname in ("HELMCoreLibrary.json", "custom_monomers.json"):
        path = root / fname
        if path.exists():
            data = json.loads(path.read_text())
            if isinstance(data, list):
                entries.extend(data)

    seen = set()
    row = 2
    for e in entries:
        sym = e.get("symbol") or e.get("id", "")
        pt  = e.get("polymerType", "PEPTIDE")
        if sym and sym not in seen:
            ws.cell(row=row, column=1, value=sym)
            ws.cell(row=row, column=2, value=pt)
            seen.add(sym)
            row += 1


def _write_helm_sheet(wb):
    ws = wb.create_sheet("HELM")
    for col, (h, w) in enumerate([("Name", 22), ("HELM", 80), ("Status", 16)], 1):
        _header(ws, col, 1, h, GREY_FILL)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"


def main(argv=None):
    p = argparse.ArgumentParser()
    p.add_argument("--out", default="peptide_design_template.xlsx")
    args = p.parse_args(argv)

    wb = openpyxl.Workbook()
    ws_design = wb.active
    ws_design.title = "Design"

    headers = _build_design_headers(ws_design)
    _write_example_rows(ws_design, headers)
    _write_monomer_db_sheet(wb)
    _write_helm_sheet(wb)

    # Example sheet: copy of Design (read-only reference)
    ws_ex = wb.create_sheet("Example")
    for row in ws_design.iter_rows():
        for cell in row:
            ws_ex.cell(row=cell.row, column=cell.column, value=cell.value)
    ws_ex.freeze_panes = "B2"

    out = Path(args.out)
    wb.save(out)
    print(f"Template written -> {out}")
    print("Next: open in Excel, import helm_builder.bas via VBE (Alt+F11 -> File -> Import)")


if __name__ == "__main__":
    main()
